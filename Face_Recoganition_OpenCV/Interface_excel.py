import sys
import os
from MatchData import matchData
from SaveData import saveData
from LoadData import loadData
from camera import Camera
from interFace_msg import message
from openpyxl import Workbook, load_workbook
from datetime import datetime

class interFace:
    def __init__(self, confidence_match=0.3, confidence_save=0.3, showConfidence=True):
        self.path = os.path.join(os.getcwd(), 'data')
        self.load_data = loadData()
        self.confidence_match = confidence_match
        self.save_confidence = confidence_save
        self.showConfidence = showConfidence
        self.run()

    def run(self):
        wb = self.create_excel_file(self.path)
        ws = wb.active
        while True:
            print("Welcome to the face recognition system")
            print("1. Register a new user via Camera")
            print("2. Register a new user via Image")
            print("3. Start Recognition")
            print("4. Check if User Data exist or not.")
            print("5. Configure Confidence Levels (Default 60%)")
            print("6. Exit")
            try:
                choice = int(input("Enter your choice: "))
            except ValueError:
                print("Invalid input. Please enter a number.")
                input("Press any key to continue...")
                self.clear_screen()
                continue

            if choice == 1:
                self.registerViaCamera()
            elif choice == 2:
                imag = input("Enter the path of the image: ")
                self.registerViaImage(imag)
            elif choice == 3:
                self.recognize(ws,wb)
            elif choice == 4:
                if self.is_valid_path(os.path.join(self.path,f'{input("Enter User Name ")}_{input("Enter User ID ")}_0.npy')):
                    print("The User Data Exists. ")
                else:
                    print("The User Data does not exist.")
                print()
                input("Press any key to continue...")
                self.clear_screen()
            elif choice==5:
                self.configureConfidence()
            elif choice == 6:
                print("Exiting program...")
                sys.exit()
            else:
                print("Invalid choice")
                input("Press any key to continue...")
                self.clear_screen()

    def registerViaCamera(self):
        if Camera().isSaved:
            self.registerViaImage(os.path.join(self.path, 'temp.jpg'))
        else:
            self.clear_screen()

    def registerViaImage(self, image_path):
        if not os.path.exists(image_path):
            message("Image file does not exist. Please provide a valid image path.", input_key=True)
            return
        saveStatus = saveData(image_path,self.load_data,self.showConfidence,self.save_confidence)
        if saveStatus.flag:
            print("User registered successfully")
            self.load_data = loadData()
        message("", input_key=True)
        return


    def create_excel_file(self,path):
        if not os.path.exists(path):
            os.makedirs(path)
        if not os.path.exists(os.path.join(self.path, 'data.xlsx')):
            wb = Workbook()
            sheet = wb.active
            sheet.title = "User Data"
            sheet.append(["Name", "ID", "Confidence", "Date", "Time"])
            wb.save(os.path.join(self.path, 'data.xlsx'))
        wb = load_workbook(os.path.join(self.path, 'data.xlsx'))
        return wb

    def recognize(self,ws,wb):
        print("Starting Recognition... Press 'q' to quit.")
        while True:
            cam = Camera()
            cam.capture()
            if cam.isSaved:
                cropped_face, cropped_face_locations = Camera.crop_face(os.path.join(self.path, 'temp.jpg'))
                for cropped_face, location in zip(cropped_face, cropped_face_locations):
                    cropped_face_path = os.path.join(self.path, 'cropped.jpg')
                    Camera.img_write(cropped_face, cropped_face_path)
                    matcher = matchData(cropped_face_path, load_data=self.load_data)
                    matched = matcher.result
                    if matched is not None and matched[3] < self.confidence_match:
                        ws.append([matched[0], matched[1], str((1-matched[3])*100), datetime.now().strftime("%d/%m/%Y"), datetime.now().strftime("%H:%M:%S")])   
                        wb.save(os.path.join(self.path, 'data.xlsx')) 
            else:
                cam.destroy()
                break

    def configureConfidence(self):
        while True:
            self.clear_screen()
            print("Select the item you want to configure.")
            print("1.Change Whether to Show Confidence or not.")
            print("2.Change Confidence for Recoganising (Matching)")
            print("3.Change Confidence for Saving (Registering)")
            print("4.Go Back to Main Menu")

            choice = input("Enter your choice: ")
            if choice == '1':
                self.showConfidence = not self.showConfidence
                print(f"Show Confidence is now set to {self.showConfidence}")
            elif choice == '2':
                try:
                    confidence = float(input("Enter the new confidence level for recognition in percentage: "))
                    if 0 <= confidence <= 100:
                        self.confidence_match = (100-confidence)/100
                        print(f"Recognition confidence set to {confidence:.2f}%")
                    else:
                        print("Invalid confidence level. Please enter a value between 0 and 100.")
                except ValueError:
                    print("Invalid input. Please enter a number.")
            elif choice == '3':
                try:
                    confidence = float(input("Enter the new confidence level for saving in percentage: "))
                    if 0 <= confidence <= 100:
                        self.save_confidence = (100-confidence)/100
                        print(f"Saving confidence set to {confidence:.2f}%")
                    else:
                        print("Invalid confidence level. Please enter a value between 0 and 100.")
                except ValueError:
                    print("Invalid input. Please enter a number.")
            elif choice == '4':
                return
            else:
                print("Invalid choice")
            input("Press any key to continue...")

    def is_valid_path(self, path):
        return os.path.exists(path) and os.path.isfile(path)
    
    def clear_screen(self):
        os.system('cls' if os.name == 'nt' else 'clear')

    
            
