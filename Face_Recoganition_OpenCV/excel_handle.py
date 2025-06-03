from openpyxl import Workbook, load_workbook , utils
from openpyxl.worksheet.dimensions import ColumnDimension
from datetime import datetime
import time
import sys
import os

class Excel_handle:
    def __init__(self,path):
        self.path = path
        self.wb = self.create_excel_file()
        if self.wb is None:
            print("Failed to create or load the Excel file. Exiting...")
            sys.exit(1)
        self.ws = self.wb.active

    def create_excel_file(self):   
        if not os.path.exists(self.path):
            wb = Workbook()
            ws = wb.active
            ws.title = "User Data"
            ws.append(["ID", "Name", "Confidence","Count"])
            ws.column_dimensions['A'].width = len("ID")+3
            ws.column_dimensions['B'].width = len("Name")+3
            ws.column_dimensions['C'].width = len("Confidence")+3
            ws.column_dimensions['D'].width = len("Count")+3
            wb.save(self.path)
            wb.close()
        try:
            wb = load_workbook(self.path)
        except Exception as e:
            print(f"Error loading Excel File: {e}")
            print("Do you want to delete the existing file and create a new one? (y/n)")
            choice = input().strip().lower()
            if choice == 'y':
                time.sleep(1)
                try:
                    os.remove(self.path)
                except PermissionError as e:
                    print(f"Permission Error: {e}. Please close the file and try again.")
                    return None
                return self.create_excel_file()
            else:
                return None
        return wb
    
    def get_row_number(self, id):
        for rn in range(2, self.ws.max_row + 1):  # Start from row 2 (skip header)
            cell_value = self.ws.cell(row=rn, column=1).value  # ID is in column 1
            if cell_value == id:
                return rn
        return None
    

    def write_to_excel(self,name, id, confidence,tg,dt = None):
        row_num = self.get_row_number(id)
        if dt is None:
            dt = datetime.now().strftime("%d/%m/%Y") + " - " + datetime.now().strftime("%H:%M:%S")
        confidence = round((1-confidence) * 100, 2)  # Convert to percentage
        if row_num is None:
            self.ws.append([id, name, confidence, 1,dt])
            row_num = self.ws.max_row
            
        count = self.ws.cell(row=row_num, column=4).value 
        if not isinstance(count,int):
            count = 0   
        previous_time = self.ws.cell(row=row_num, column=4+count).value
        if isinstance(previous_time, str) and self.timeGapInSeconds(previous_time, dt) < tg:
            print("Duplicate entry detected. Skipping...")
            return
        existing_conf = self.ws.cell(row=row_num, column=3).value
        if not isinstance(existing_conf, (int, float)):
            existing_conf = confidence
        self.ws.cell(row=row_num, column=3).value = min(existing_conf, confidence)
        self.ws.cell(row=row_num, column=4).value = count + 1
        self.ws.cell(row=row_num, column=5+count).value = dt

        for col, val in zip(['A', 'B', 'C', 'D', utils.get_column_letter(5 + count)], [str(id), name, str(confidence), str(count + 1), dt]):
            width = len(str(val)) + 3
            dim = self.ws.column_dimensions.get(col)
            if dim is None:
                self.ws.column_dimensions[col] = ColumnDimension(self.ws, col)
            self.ws.column_dimensions[col].width = max(self.ws.column_dimensions[col].width or 0, width)

        
        self.wb.save(self.path)
            
    def timeGapInSeconds(self, old_time, new_time):
        try:
            old_dt = datetime.strptime(old_time, "%d/%m/%Y - %H:%M:%S")
            new_dt = datetime.strptime(new_time, "%d/%m/%Y - %H:%M:%S")
            return (new_dt - old_dt).total_seconds()
        except Exception as e:
            print(f"Error in timeGapInSeconds: {e}")
            return float('inf')  # Treat invalid comparison as very large gap
    
    def read_excel(self,row,column):
        return self.ws.cell(row,column).value;

    def get_date_time_now(self):
        return datetime.now().strftime("%d/%m/%Y") + " - " + datetime.now().strftime("%H:%M:%S")
    
    def write_excel(self,row,column,value):
        self.ws.cell(row,column).value = value
        self.wb.save(self.path)


    def is_valid_time(self, time_str):
        try:
            datetime.strptime(time_str, "%d/%m/%Y - %H:%M:%S")
            return True
        except ValueError:
            return False
        
    def is_valid_date(self, date_str):
        try:
            datetime.strptime(date_str, "%d/%m/%Y")
            return True
        except ValueError:
            return False

    def get_entries_by_time_range(self,c_row,initial_time,final_time):
        entries = []
        if not self.is_valid_time(initial_time) or not self.is_valid_time(final_time):
            print("Invalid time format. Please use 'dd/mm/yyyy - hh:mm:ss'.")
            return entries
        for cl in range(5,self.ws.max_column+1):
            ex_val = self.read_excel(c_row,cl)
            if(ex_val < initial_time):
                continue
            elif(ex_val >= final_time):
                break
            else:
                entries.append(ex_val)
        return entries    
    
    def increment_entry_count(self, row_num, val = 1):
        if row_num is not None:
            count = self.ws.cell(row=row_num, column=4).value
            if not isinstance(count, int):
                count = 0
            self.ws.cell(row=row_num, column=4).value = count + val
            self.wb.save(self.path)
            return count + val
        return None